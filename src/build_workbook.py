"""
Build Excel workbook for AI Market Intelligence Dashboard
"""
import os
import sys
import pandas as pd
from datetime import datetime
from typing import Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

# Add src to path for imports
sys.path.append(os.path.dirname(os.path.abspath(__file__)))


class WorkbookBuilder:
    """Builds Excel workbook from CSV data"""
    
    def __init__(self):
        self.data_dir = 'data'
        self.output_file = 'dashboard_v1.xlsx'
        
    def load_csv_data(self) -> Dict[str, pd.DataFrame]:
        """Load data from CSV files"""
        data = {}
        
        csv_files = {
            'market_dynamics': 'market_dynamics.csv',
            'hyperscaler_metrics': 'hyperscaler_metrics.csv',
            'context_signals': 'context_signals.csv'
        }
        
        for key, filename in csv_files.items():
            filepath = os.path.join(self.data_dir, filename)
            if os.path.exists(filepath):
                try:
                    df = pd.read_csv(filepath)
                    data[key] = df
                    print(f"Loaded {len(df)} rows from {filename}")
                except Exception as e:
                    print(f"Error loading {filename}: {e}")
                    data[key] = pd.DataFrame()
            else:
                print(f"File not found: {filename}")
                data[key] = pd.DataFrame()
        
        return data
    
    def create_market_dynamics_sheet(self, wb: openpyxl.Workbook, df: pd.DataFrame):
        """Create Market Dynamics sheet"""
        ws = wb.create_sheet("Market_Dynamics")
        
        if df.empty:
            ws['A1'] = "No market dynamics data available"
            return
        
        # Create pivoted view
        pivot_data = self._create_market_dynamics_pivot(df)
        
        # Add title
        ws['A1'] = "AI Market Intelligence Dashboard - Market Dynamics"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Add pivot table
        start_row = 3
        self._add_pivot_table(ws, pivot_data, start_row, "Market Dynamics by Vertical")
        
        # Add startup churn ratio table
        churn_start_row = start_row + len(pivot_data) + 3
        self._add_churn_ratio_table(ws, df, churn_start_row)
        
        # Style the sheet
        self._style_market_dynamics_sheet(ws)
    
    def create_hyperscalers_sheet(self, wb: openpyxl.Workbook, df: pd.DataFrame):
        """Create Hyperscalers sheet"""
        ws = wb.create_sheet("Hyperscalers")
        
        if df.empty:
            ws['A1'] = "No hyperscaler metrics data available"
            return
        
        # Add title
        ws['A1'] = "AI Market Intelligence Dashboard - Hyperscaler Metrics"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:J1')
        
        # Create hyperscaler summary table
        summary_data = self._create_hyperscaler_summary(df)
        
        # Add summary table
        start_row = 3
        self._add_hyperscaler_table(ws, summary_data, start_row)
        
        # Add charts
        chart_start_row = start_row + len(summary_data) + 3
        self._add_hyperscaler_charts(ws, summary_data, chart_start_row)
        
        # Style the sheet
        self._style_hyperscalers_sheet(ws)
    
    def create_context_sheet(self, wb: openpyxl.Workbook, df: pd.DataFrame):
        """Create Context sheet"""
        ws = wb.create_sheet("Context")
        
        if df.empty:
            ws['A1'] = "No context signals data available"
            return
        
        # Add title
        ws['A1'] = "AI Market Intelligence Dashboard - Context Signals"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Add full context table
        start_row = 3
        self._add_context_table(ws, df, start_row)
        
        # Add top-3 summary blocks
        summary_start_row = start_row + len(df) + 3
        self._add_context_summaries(ws, df, summary_start_row)
        
        # Style the sheet
        self._style_context_sheet(ws)
    
    def _create_market_dynamics_pivot(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create pivoted view of market dynamics"""
        # Filter for main metrics (exclude derived ones for now)
        main_metrics = ['funding_total_usd', 'deals_count', 'startups_new', 'startups_funded', 'shutdowns', 'acquisitions']
        main_df = df[df['metric'].isin(main_metrics)]
        
        # Create pivot table
        pivot = main_df.pivot_table(
            index='metric',
            columns='vertical',
            values='value',
            aggfunc='first',
            fill_value=0
        )
        
        # Add derived metrics
        derived_metrics = df[df['metric'].isin(['avg_deal_size_usd', 'startup_churn_ratio'])]
        for _, row in derived_metrics.iterrows():
            if row['metric'] not in pivot.index:
                pivot.loc[row['metric']] = 0
            pivot.loc[row['metric'], row['vertical']] = row['value']
        
        return pivot
    
    def _create_hyperscaler_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create summary table for hyperscaler metrics"""
        # Filter for main metrics
        main_metrics = ['announcements_count', 'initiatives_new', 'initiatives_total_cume', 'initiative_momentum_pct']
        main_df = df[df['metric'].isin(main_metrics)]
        
        # Create summary table
        summary = main_df.pivot_table(
            index=['hyperscaler', 'vertical'],
            columns='metric',
            values='value',
            aggfunc='first',
            fill_value=0
        ).reset_index()
        
        return summary
    
    def _add_pivot_table(self, ws: openpyxl.Workbook.Worksheet, pivot_df: pd.DataFrame, start_row: int, title: str):
        """Add pivot table to worksheet"""
        # Add title
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(size=14, bold=True)
        
        # Add headers
        header_row = start_row + 1
        ws.cell(row=header_row, column=1, value="Metric")
        ws.cell(row=header_row, column=1).font = Font(bold=True)
        
        col = 2
        for vertical in pivot_df.columns:
            ws.cell(row=header_row, column=col, value=vertical.title())
            ws.cell(row=header_row, column=col).font = Font(bold=True)
            col += 1
        
        # Add data
        row = header_row + 1
        for metric in pivot_df.index:
            ws.cell(row=row, column=1, value=metric.replace('_', ' ').title())
            col = 2
            for vertical in pivot_df.columns:
                value = pivot_df.loc[metric, vertical]
                if pd.notna(value) and value != 0:
                    ws.cell(row=row, column=col, value=value)
                col += 1
            row += 1
    
    def _add_churn_ratio_table(self, ws: openpyxl.Workbook.Worksheet, df: pd.DataFrame, start_row: int):
        """Add startup churn ratio table"""
        # Add title
        ws.cell(row=start_row, column=1, value="Startup Churn Ratio by Vertical")
        ws.cell(row=start_row, column=1).font = Font(size=14, bold=True)
        
        # Add headers
        header_row = start_row + 1
        ws.cell(row=header_row, column=1, value="Vertical")
        ws.cell(row=header_row, column=2, value="Churn Ratio")
        ws.cell(row=header_row, column=1).font = Font(bold=True)
        ws.cell(row=header_row, column=2).font = Font(bold=True)
        
        # Add data
        churn_data = df[df['metric'] == 'startup_churn_ratio']
        row = header_row + 1
        for _, data_row in churn_data.iterrows():
            ws.cell(row=row, column=1, value=data_row['vertical'].title())
            ws.cell(row=row, column=2, value=data_row['value'])
            row += 1
    
    def _add_hyperscaler_table(self, ws: openpyxl.Workbook.Worksheet, summary_df: pd.DataFrame, start_row: int):
        """Add hyperscaler summary table"""
        # Add title
        ws.cell(row=start_row, column=1, value="Hyperscaler Metrics Summary")
        ws.cell(row=start_row, column=1).font = Font(size=14, bold=True)
        
        # Add headers
        header_row = start_row + 1
        headers = ['Hyperscaler', 'Vertical', 'Announcements', 'New Initiatives', 'Total Initiatives', 'Momentum %']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=header_row, column=col, value=header)
            ws.cell(row=header_row, column=col).font = Font(bold=True)
        
        # Add data
        row = header_row + 1
        for _, data_row in summary_df.iterrows():
            ws.cell(row=row, column=1, value=data_row['hyperscaler'].title())
            ws.cell(row=row, column=2, value=data_row['vertical'].title())
            ws.cell(row=row, column=3, value=data_row.get('announcements_count', 0))
            ws.cell(row=row, column=4, value=data_row.get('initiatives_new', 0))
            ws.cell(row=row, column=5, value=data_row.get('initiatives_total_cume', 0))
            ws.cell(row=row, column=6, value=data_row.get('initiative_momentum_pct', 0))
            row += 1
    
    def _add_hyperscaler_charts(self, ws: openpyxl.Workbook.Worksheet, summary_df: pd.DataFrame, start_row: int):
        """Add charts for hyperscaler metrics"""
        # Create a simple bar chart for initiatives
        chart = BarChart()
        chart.title = "New Initiatives by Hyperscaler and Vertical"
        chart.x_axis.title = "Hyperscaler-Vertical"
        chart.y_axis.title = "Count"
        
        # Add chart data (simplified)
        data_start_row = start_row + 1
        ws.cell(row=data_start_row, column=1, value="Chart data would go here")
        ws.cell(row=data_start_row, column=1).font = Font(italic=True)
    
    def _add_context_table(self, ws: openpyxl.Workbook.Worksheet, df: pd.DataFrame, start_row: int):
        """Add full context signals table"""
        # Add title
        ws.cell(row=start_row, column=1, value="Context Signals")
        ws.cell(row=start_row, column=1).font = Font(size=14, bold=True)
        
        # Add headers
        header_row = start_row + 1
        headers = ['Date', 'Type', 'Title', 'Summary', 'Vertical', 'Sentiment', 'Source']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=header_row, column=col, value=header)
            ws.cell(row=header_row, column=col).font = Font(bold=True)
        
        # Add data (limit to first 50 rows for readability)
        row = header_row + 1
        for _, data_row in df.head(50).iterrows():
            ws.cell(row=row, column=1, value=data_row['date'])
            ws.cell(row=row, column=2, value=data_row['signal_type'])
            ws.cell(row=row, column=3, value=data_row['title'])
            ws.cell(row=row, column=4, value=data_row['summary_140'])
            ws.cell(row=row, column=5, value=data_row['vertical'])
            ws.cell(row=row, column=6, value=data_row['sentiment'])
            ws.cell(row=row, column=7, value=data_row['source_name'])
            row += 1
    
    def _add_context_summaries(self, ws: openpyxl.Workbook.Worksheet, df: pd.DataFrame, start_row: int):
        """Add top-3 summaries by signal type"""
        # Add title
        ws.cell(row=start_row, column=1, value="Top 3 Most Recent Signals by Type")
        ws.cell(row=start_row, column=1).font = Font(size=14, bold=True)
        
        signal_types = ['policy', 'news', 'adoption_or_risk_signal']
        current_row = start_row + 1
        
        for signal_type in signal_types:
            # Add signal type header
            ws.cell(row=current_row, column=1, value=f"{signal_type.title()} Signals")
            ws.cell(row=current_row, column=1).font = Font(size=12, bold=True)
            current_row += 1
            
            # Get top 3 for this signal type
            type_data = df[df['signal_type'] == signal_type].head(3)
            
            for _, data_row in type_data.iterrows():
                ws.cell(row=current_row, column=1, value=f"• {data_row['title']}")
                ws.cell(row=current_row, column=2, value=data_row['date'])
                current_row += 1
            
            current_row += 1  # Add spacing
    
    def _style_market_dynamics_sheet(self, ws: openpyxl.Workbook.Worksheet):
        """Apply styling to market dynamics sheet"""
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Apply header styling
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.isupper():
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
    
    def _style_hyperscalers_sheet(self, ws: openpyxl.Workbook.Worksheet):
        """Apply styling to hyperscalers sheet"""
        # Similar styling as market dynamics
        self._style_market_dynamics_sheet(ws)
    
    def _style_context_sheet(self, ws: openpyxl.Workbook.Worksheet):
        """Apply styling to context sheet"""
        # Similar styling as market dynamics
        self._style_market_dynamics_sheet(ws)
    
    def build_workbook(self):
        """Build the complete Excel workbook"""
        print("Building Excel workbook...")
        
        # Load data
        data = self.load_csv_data()
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        if 'market_dynamics' in data:
            self.create_market_dynamics_sheet(wb, data['market_dynamics'])
        
        if 'hyperscaler_metrics' in data:
            self.create_hyperscalers_sheet(wb, data['hyperscaler_metrics'])
        
        if 'context_signals' in data:
            self.create_context_sheet(wb, data['context_signals'])
        
        # Save workbook
        output_path = os.path.join(self.data_dir, self.output_file)
        wb.save(output_path)
        print(f"Workbook saved to {output_path}")
        
        return output_path


def main():
    """Main function to build Excel workbook"""
    print("Starting Excel workbook creation...")
    
    builder = WorkbookBuilder()
    output_path = builder.build_workbook()
    
    print(f"Excel workbook creation complete: {output_path}")


if __name__ == "__main__":
    main()
